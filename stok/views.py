from datetime import date, datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .forms import (
    DemoCihazForm,
    ExcelIceriAktarForm,
    FirmaGonderiForm,
    SteriSenseForm,
    StokMiktarForm,
    TopluUrunForm,
    UrunForm,
)

from .models import (
    DemoCihaz,
    FirmaGonderi,
    SteriSenseKaydi,
    StokHareketi,
    Urun,
)


def _metin(deger):
    if deger is None:
        return ""
    return str(deger).strip()


def _excel_tarihi(deger):
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    if isinstance(deger, str):
        temiz = deger.strip()
        for bicim in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(temiz, bicim).date()
            except ValueError:
                pass
    return None


def _baslik_indeksleri(sayfa):
    return {
        _metin(hucre.value).casefold(): indeks
        for indeks, hucre in enumerate(sayfa[1])
        if _metin(hucre.value)
    }


def _deger(satir, basliklar, *adlar):
    for ad in adlar:
        indeks = basliklar.get(ad.casefold())
        if indeks is not None and indeks < len(satir):
            return satir[indeks]
    return None


def _excel_verilerini_aktar(dosya):
    kitap = load_workbook(dosya, data_only=True, read_only=True)
    sonuc = {
        "stok_eklendi": 0,
        "firma_eklendi": 0,
        "demo_eklendi": 0,
        "sterisense_eklendi": 0,
        "atlandi": 0,
        "hata": 0,
    }

    with transaction.atomic():
        if "Stoklar" in kitap.sheetnames:
            sayfa = kitap["Stoklar"]
            basliklar = _baslik_indeksleri(sayfa)
            for satir in sayfa.iter_rows(min_row=2, values_only=True):
                try:
                    urun_adi = _metin(_deger(satir, basliklar, "Ürün Adı", "Ürün"))
                    seri_no = _metin(_deger(satir, basliklar, "Seri No", "Seri Numarası"))
                    if not urun_adi or not seri_no:
                        sonuc["atlandi"] += 1
                        continue
                    if Urun.objects.filter(seri_no__iexact=seri_no).exists():
                        sonuc["atlandi"] += 1
                        continue
                    giris_tarihi = _excel_tarihi(_deger(satir, basliklar, "Giriş Tarihi")) or date.today()
                    adet_ham = _deger(satir, basliklar, "Adet")
                    try:
                        adet = max(0, int(adet_ham if adet_ham not in (None, "") else 1))
                    except (TypeError, ValueError):
                        adet = 1
                    urun = Urun.objects.create(
                        urun_adi=urun_adi,
                        seri_no=seri_no,
                        giris_tarihi=giris_tarihi,
                        adet=adet,
                    )
                    StokHareketi.objects.create(
                        urun=urun,
                        islem_turu="giris",
                        miktar=adet,
                        aciklama="Excel dosyasından içe aktarıldı.",
                    )
                    sonuc["stok_eklendi"] += 1
                except Exception:
                    sonuc["hata"] += 1

        if "Firma Gönderileri" in kitap.sheetnames:
            sayfa = kitap["Firma Gönderileri"]
            basliklar = _baslik_indeksleri(sayfa)
            for satir in sayfa.iter_rows(min_row=2, values_only=True):
                try:
                    firma = _metin(_deger(satir, basliklar, "Firma", "Firma Adı"))
                    urun_adi = _metin(_deger(satir, basliklar, "Ürün Adı", "Ürün"))
                    seri_no = _metin(_deger(satir, basliklar, "Seri Numarası", "Seri No"))
                    if not firma or not urun_adi:
                        sonuc["atlandi"] += 1
                        continue
                    tarih = (
                        _excel_tarihi(_deger(satir, basliklar, "Çıkış Tarihi", "Gönderim Tarihi"))
                        or _excel_tarihi(_deger(satir, basliklar, "Giriş Tarihi"))
                        or date.today()
                    )
                    notlar = _metin(_deger(satir, basliklar, "Notlar", "Not"))
                    if FirmaGonderi.objects.filter(
                        firma_adi__iexact=firma,
                        urun_adi__iexact=urun_adi,
                        seri_no__iexact=seri_no,
                        gonderim_tarihi=tarih,
                    ).exists():
                        sonuc["atlandi"] += 1
                        continue
                    FirmaGonderi.objects.create(
                        firma_adi=firma,
                        urun_adi=urun_adi,
                        seri_no=seri_no,
                        gonderim_tarihi=tarih,
                        notlar=notlar,
                        aktif=True,
                    )
                    sonuc["firma_eklendi"] += 1
                except Exception:
                    sonuc["hata"] += 1

        if "Demo Cihazlar" in kitap.sheetnames:
            sayfa = kitap["Demo Cihazlar"]
            basliklar = _baslik_indeksleri(sayfa)
            for satir in sayfa.iter_rows(min_row=2, values_only=True):
                try:
                    urun_adi = _metin(_deger(satir, basliklar, "Ürün Adı", "Ürün"))
                    seri_no = _metin(_deger(satir, basliklar, "Seri Numarası", "Seri No"))
                    if not urun_adi or not seri_no:
                        sonuc["atlandi"] += 1
                        continue
                    if DemoCihaz.objects.filter(urun_adi__iexact=urun_adi, seri_no__iexact=seri_no).exists():
                        sonuc["atlandi"] += 1
                        continue
                    DemoCihaz.objects.create(urun_adi=urun_adi, seri_no=seri_no)
                    sonuc["demo_eklendi"] += 1
                except Exception:
                    sonuc["hata"] += 1

        if "Sterisense" in kitap.sheetnames:
            sayfa = kitap["Sterisense"]
            basliklar = _baslik_indeksleri(sayfa)
            for satir in sayfa.iter_rows(min_row=2, values_only=True):
                try:
                    bayi = _metin(_deger(satir, basliklar, "Bayi"))
                    firma = _metin(_deger(satir, basliklar, "Firma"))
                    urun_adi = _metin(_deger(satir, basliklar, "Ürün", "Ürün Adı"))
                    seri_no = _metin(_deger(satir, basliklar, "Seri No", "Seri Numarası"))
                    if not bayi or not firma or not urun_adi or not seri_no:
                        sonuc["atlandi"] += 1
                        continue
                    tarih = _excel_tarihi(_deger(satir, basliklar, "Tarih")) or date.today()
                    notlar = _metin(_deger(satir, basliklar, "Notlar", "Not"))
                    if SteriSenseKaydi.objects.filter(
                        bayi__iexact=bayi,
                        firma__iexact=firma,
                        urun_adi__iexact=urun_adi,
                        seri_no__iexact=seri_no,
                        tarih=tarih,
                    ).exists():
                        sonuc["atlandi"] += 1
                        continue
                    SteriSenseKaydi.objects.create(
                        bayi=bayi,
                        firma=firma,
                        urun_adi=urun_adi,
                        seri_no=seri_no,
                        tarih=tarih,
                        notlar=notlar,
                    )
                    sonuc["sterisense_eklendi"] += 1
                except Exception:
                    sonuc["hata"] += 1

    kitap.close()
    return sonuc


@login_required
def excel_ice_aktar(request):
    form = ExcelIceriAktarForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        try:
            sonuc = _excel_verilerini_aktar(form.cleaned_data["excel_dosyasi"])
            messages.success(
                request,
                (
                    f"Aktarım tamamlandı: {sonuc['stok_eklendi']} stok, "
                    f"{sonuc['firma_eklendi']} firma gönderisi, "
                    f"{sonuc['demo_eklendi']} demo, "
                    f"{sonuc['sterisense_eklendi']} SteriSense kaydı eklendi. "
                    f"{sonuc['atlandi']} satır atlandı, {sonuc['hata']} satırda hata oluştu."
                ),
            )
            return redirect("stok:urun_listesi")
        except Exception as hata:
            form.add_error("excel_dosyasi", f"Dosya aktarılamadı: {hata}")

    return render(
        request,
        "stok/excel_ice_aktar.html",
        {"form": form, "aktif_menu": "stoklar"},
    )


@login_required
def urun_listesi(request):
    arama = request.GET.get("q", "").strip()

    urunler = Urun.objects.all().order_by(
        "urun_adi",
        "giris_tarihi",
        "seri_no",
    )

    if arama:
        urunler = urunler.filter(
            Q(urun_adi__icontains=arama)
            | Q(seri_no__icontains=arama)
        )

    grup_sozlugu = {}

    for urun in urunler:
        anahtar = urun.urun_adi.strip().casefold()

        if anahtar not in grup_sozlugu:
            grup_sozlugu[anahtar] = {
                "urun_adi": urun.urun_adi,
                "toplam_adet": 0,
                "kayitlar": [],
            }

        grup_sozlugu[anahtar]["toplam_adet"] += urun.adet
        grup_sozlugu[anahtar]["kayitlar"].append(urun)

    urun_gruplari = list(grup_sozlugu.values())

    return render(
        request,
        "stok/urun_listesi.html",
        {
            "urun_gruplari": urun_gruplari,
            "arama": arama,
            "aktif_menu": "stoklar",
        },
    )


@login_required
def urun_ekle(request):
    form = TopluUrunForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        urun_adi = form.cleaned_data["urun_adi"].strip()
        giris_tarihi = form.cleaned_data["giris_tarihi"]
        seri_numaralari = form.cleaned_data["seri_numaralari"]

        eklenen = 0
        atlanan = 0

        with transaction.atomic():
            for seri_no in seri_numaralari:
                if Urun.objects.filter(
                    seri_no__iexact=seri_no
                ).exists():
                    atlanan += 1
                    continue

                urun = Urun.objects.create(
                    urun_adi=urun_adi,
                    seri_no=seri_no,
                    giris_tarihi=giris_tarihi,
                    adet=1,
                )

                StokHareketi.objects.create(
                    urun=urun,
                    islem_turu="giris",
                    miktar=1,
                    aciklama="Toplu seri numarası kaydıyla eklendi.",
                )

                eklenen += 1

        if eklenen:
            messages.success(
                request,
                f"{eklenen} seri numarası stoğa eklendi."
            )

        if atlanan:
            messages.warning(
                request,
                f"{atlanan} seri numarası daha önce kayıtlı olduğu için atlandı."
            )

        return redirect("stok:urun_listesi")

    return render(
        request,
        "stok/toplu_urun_formu.html",
        {
            "form": form,
            "sayfa_basligi": "YENİ ÜRÜN EKLE",
            "aktif_menu": "stoklar",
        },
    )


@login_required
def urun_duzenle(request, urun_id):
    urun = get_object_or_404(Urun, pk=urun_id)
    onceki_adet = urun.adet
    form = UrunForm(request.POST or None, instance=urun)

    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            urun = form.save()
            fark = urun.adet - onceki_adet
            StokHareketi.objects.create(
                urun=urun,
                islem_turu="duzenleme",
                miktar=fark,
                aciklama="Ürün bilgileri düzenlendi.",
            )

        messages.success(request, "Ürün bilgileri güncellendi.")
        return redirect("stok:urun_listesi")

    return render(
        request,
        "stok/urun_formu.html",
        {
            "form": form,
            "sayfa_basligi": "ÜRÜN DÜZENLE",
            "aktif_menu": "stoklar",
        },
    )


@login_required
def stok_girisi(request, urun_id):
    urun = get_object_or_404(Urun, pk=urun_id)
    form = StokMiktarForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        miktar = form.cleaned_data["miktar"]
        aciklama = form.cleaned_data["aciklama"] or "Manuel stok girişi"

        with transaction.atomic():
            urun.adet += miktar
            urun.save(update_fields=["adet"])
            StokHareketi.objects.create(
                urun=urun,
                islem_turu="giris",
                miktar=miktar,
                aciklama=aciklama,
            )

        messages.success(request, f"{urun.urun_adi} stoğuna {miktar} adet eklendi.")
        return redirect("stok:urun_listesi")

    return render(
        request,
        "stok/stok_miktar_formu.html",
        {
            "form": form,
            "urun": urun,
            "sayfa_basligi": "STOK GİRİŞİ",
            "buton_metni": "Kaydet",
            "aktif_menu": "stoklar",
        },
    )


@login_required
def stok_cikisi(request, urun_id):
    urun = get_object_or_404(Urun, pk=urun_id)
    form = StokMiktarForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        miktar = form.cleaned_data["miktar"]

        if miktar > urun.adet:
            form.add_error("miktar", f"Mevcut stok miktarı: {urun.adet}")
        else:
            aciklama = form.cleaned_data["aciklama"] or "Manuel stok çıkışı"

            with transaction.atomic():
                urun.adet -= miktar
                urun.save(update_fields=["adet"])
                StokHareketi.objects.create(
                    urun=urun,
                    islem_turu="cikis",
                    miktar=-miktar,
                    aciklama=aciklama,
                )

            messages.success(request, f"{urun.urun_adi} stoğundan {miktar} adet düşüldü.")
            return redirect("stok:urun_listesi")

    return render(
        request,
        "stok/stok_miktar_formu.html",
        {
            "form": form,
            "urun": urun,
            "sayfa_basligi": "STOK ÇIKIŞI",
            "buton_metni": "Kaydet",
            "aktif_menu": "stoklar",
        },
    )


@login_required
def hareketler(request, urun_id):
    urun = get_object_or_404(Urun, pk=urun_id)
    return render(
        request,
        "stok/hareketler.html",
        {
            "urun": urun,
            "hareketler": urun.hareketler.all(),
            "aktif_menu": "stoklar",
        },
    )


@login_required
@require_POST
def urun_sil(request, urun_id):
    urun = get_object_or_404(Urun, pk=urun_id)
    urun_adi = urun.urun_adi
    urun.delete()
    messages.success(request, f"{urun_adi} silindi.")
    return redirect("stok:urun_listesi")


@login_required
def excel_aktar(request):
    workbook = Workbook()
    sayfa = workbook.active
    sayfa.title = "Stoklar"
    sayfa.append(["Ürün Adı", "Seri No", "Giriş Tarihi", "Adet"])

    for hucre in sayfa[1]:
        hucre.font = Font(bold=True)

    for urun in Urun.objects.all().order_by("urun_adi", "seri_no"):
        sayfa.append([urun.urun_adi, urun.seri_no, urun.giris_tarihi, urun.adet])

    for sutun in sayfa.columns:
        uzunluk = max(len(str(hucre.value or "")) for hucre in sutun) + 2
        sayfa.column_dimensions[get_column_letter(sutun[0].column)].width = uzunluk

    dosya = BytesIO()
    workbook.save(dosya)
    dosya.seek(0)

    response = HttpResponse(
        dosya.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Stoklar.xlsx"'
    return response


# --------------------------------------------------
# FİRMA GÖNDERİLERİ
# --------------------------------------------------

@login_required
@login_required
def firma_gonderileri(request):
    arama = request.GET.get("q", "").strip()

    gonderiler = FirmaGonderi.objects.filter(
        aktif=True
    ).order_by(
        "firma_adi",
        "-gonderim_tarihi",
        "-id",
    )

    if arama:
        gonderiler = gonderiler.filter(
            Q(firma_adi__icontains=arama)
            | Q(urun_adi__icontains=arama)
            | Q(seri_no__icontains=arama)
            | Q(notlar__icontains=arama)
        )

    firma_sozlugu = {}

    for gonderi in gonderiler:
        anahtar = gonderi.firma_adi.strip().casefold()

        if anahtar not in firma_sozlugu:
            firma_sozlugu[anahtar] = {
                "firma_adi": gonderi.firma_adi,
                "toplam_gonderi": 0,
                "son_gonderim_tarihi": gonderi.gonderim_tarihi,
                "gonderiler": [],
            }

        firma_sozlugu[anahtar]["toplam_gonderi"] += 1
        firma_sozlugu[anahtar]["gonderiler"].append(gonderi)

        if (
            gonderi.gonderim_tarihi
            > firma_sozlugu[anahtar]["son_gonderim_tarihi"]
        ):
            firma_sozlugu[anahtar]["son_gonderim_tarihi"] = (
                gonderi.gonderim_tarihi
            )

    firma_gruplari = list(firma_sozlugu.values())

    return render(
        request,
        "stok/firma_gonderileri.html",
        {
            "firma_gruplari": firma_gruplari,
            "arama": arama,
            "aktif_menu": "gonderiler",
        },
    )


@login_required
def firma_gonderi_ekle(request):
    form = FirmaGonderiForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        stok_urunu = form.cleaned_data.get("stok_urunu")
        urun_adi = form.cleaned_data.get("urun_adi", "").strip()
        seri_no = form.cleaned_data.get("seri_no", "").strip()

        # Stoktan seçim yapılmadıysa manuel girilen bilgilerle ürünü ara.
        if not stok_urunu and urun_adi and seri_no:
            stok_urunu = Urun.objects.filter(
                urun_adi__iexact=urun_adi,
                seri_no__iexact=seri_no,
                adet__gt=0,
            ).first()

        with transaction.atomic():
            gonderi = form.save(commit=False)

            # Ürün stokta bulunduysa stokla ilişkilendir ve 1 adet düş.
            if stok_urunu:
                gonderi.stok_urunu = stok_urunu
                gonderi.urun_adi = stok_urunu.urun_adi
                gonderi.seri_no = stok_urunu.seri_no

                stok_urunu.adet -= 1
                stok_urunu.save(update_fields=["adet"])

                StokHareketi.objects.create(
                    urun=stok_urunu,
                    islem_turu="firma_gonderisi",
                    miktar=-1,
                    aciklama=(
                        f"Firma: {gonderi.firma_adi}. "
                        f"{gonderi.notlar}"
                    ).strip(),
                )

            # Stokta bulunmadıysa manuel bilgilerle normal kayıt oluştur.
            else:
                gonderi.stok_urunu = None
                gonderi.urun_adi = urun_adi
                gonderi.seri_no = seri_no

            gonderi.aktif = True
            gonderi.save()

        if stok_urunu:
            messages.success(
                request,
                "Firma gönderisi kaydedildi ve ürün stoktan düşüldü.",
            )
        else:
            messages.success(
                request,
                "Firma gönderisi manuel olarak kaydedildi. Stoktan düşüm yapılmadı.",
            )

        return redirect("stok:firma_gonderileri")

    return render(
        request,
        "stok/firma_gonderi_formu.html",
        {
            "form": form,
            "sayfa_basligi": "YENİ GÖNDERİ",
            "aktif_menu": "gonderiler",
        },
    )


@login_required
def firma_gonderi_duzenle(request, gonderi_id):
    gonderi = get_object_or_404(FirmaGonderi, pk=gonderi_id, aktif=True)
    eski_stok_urunu = gonderi.stok_urunu
    form = FirmaGonderiForm(request.POST or None, instance=gonderi)

    if eski_stok_urunu:
        form.fields["stok_urunu"].queryset = Urun.objects.filter(
            Q(adet__gt=0) | Q(pk=eski_stok_urunu.pk)
        ).order_by("urun_adi", "seri_no")

    if request.method == "POST" and form.is_valid():
        yeni_stok_urunu = form.cleaned_data.get("stok_urunu")

        if yeni_stok_urunu and yeni_stok_urunu != eski_stok_urunu and yeni_stok_urunu.adet < 1:
            form.add_error("stok_urunu", "Seçilen ürün stokta bulunmuyor.")
        else:
            with transaction.atomic():
                if eski_stok_urunu != yeni_stok_urunu:
                    if eski_stok_urunu:
                        eski_stok_urunu.adet += 1
                        eski_stok_urunu.save(update_fields=["adet"])

                    if yeni_stok_urunu:
                        yeni_stok_urunu.adet -= 1
                        yeni_stok_urunu.save(update_fields=["adet"])

                gonderi = form.save(commit=False)
                if yeni_stok_urunu:
                    gonderi.urun_adi = yeni_stok_urunu.urun_adi
                    gonderi.seri_no = yeni_stok_urunu.seri_no
                gonderi.save()

            messages.success(request, "Firma gönderisi güncellendi.")
            return redirect("stok:firma_gonderileri")

    return render(
        request,
        "stok/firma_gonderi_formu.html",
        {
            "form": form,
            "sayfa_basligi": "GÖNDERİ DÜZENLE",
            "aktif_menu": "gonderiler",
        },
    )


@login_required
@require_POST
def firma_gonderi_geri_al(request, gonderi_id):
    gonderi = get_object_or_404(FirmaGonderi, pk=gonderi_id, aktif=True)

    with transaction.atomic():
        gonderi.aktif = False
        gonderi.save(update_fields=["aktif"])

        if gonderi.stok_urunu:
            urun = gonderi.stok_urunu
            urun.adet += 1
            urun.save(update_fields=["adet"])
            StokHareketi.objects.create(
                urun=urun,
                islem_turu="firma_donusu",
                miktar=1,
                aciklama=f"{gonderi.firma_adi} firmasından stoğa geri alındı.",
            )

    messages.success(request, "Gönderi geri alındı.")
    return redirect("stok:firma_gonderileri")


@login_required
@require_POST
def firma_gonderi_sil(request, gonderi_id):
    gonderi = get_object_or_404(FirmaGonderi, pk=gonderi_id)

    with transaction.atomic():
        if gonderi.aktif and gonderi.stok_urunu:
            urun = gonderi.stok_urunu
            urun.adet += 1
            urun.save(update_fields=["adet"])
            StokHareketi.objects.create(
                urun=urun,
                islem_turu="firma_donusu",
                miktar=1,
                aciklama="Firma gönderisi silindiği için ürün stoğa geri alındı.",
            )

        gonderi.delete()

    messages.success(request, "Firma gönderisi silindi.")
    return redirect("stok:firma_gonderileri")


@login_required
def firma_gonderileri_excel(request):
    workbook = Workbook()
    sayfa = workbook.active
    sayfa.title = "Firma Gönderileri"
    sayfa.append(["Firma Adı", "Ürün Adı", "Seri No", "Gönderim Tarihi", "Notlar"])

    for hucre in sayfa[1]:
        hucre.font = Font(bold=True)

    for gonderi in FirmaGonderi.objects.filter(aktif=True):
        sayfa.append([
            gonderi.firma_adi,
            gonderi.urun_adi,
            gonderi.seri_no,
            gonderi.gonderim_tarihi,
            gonderi.notlar,
        ])

    for sutun in sayfa.columns:
        uzunluk = max(len(str(hucre.value or "")) for hucre in sutun) + 2
        sayfa.column_dimensions[get_column_letter(sutun[0].column)].width = min(uzunluk, 60)

    dosya = BytesIO()
    workbook.save(dosya)
    dosya.seek(0)

    response = HttpResponse(
        dosya.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Firma_Gonderileri.xlsx"'
    return response


# --------------------------------------------------
# DEMO CİHAZLAR
# --------------------------------------------------

@login_required
def demo_cihazlar(request):
    arama = request.GET.get("q", "").strip()

    demo_listesi = DemoCihaz.objects.all().order_by(
        "urun_adi",
        "seri_no",
    )

    if arama:
        demo_listesi = demo_listesi.filter(
            Q(urun_adi__icontains=arama)
            | Q(seri_no__icontains=arama)
        )

    grup_sozlugu = {}

    for demo in demo_listesi:
        anahtar = demo.urun_adi.strip().casefold()

        if anahtar not in grup_sozlugu:
            grup_sozlugu[anahtar] = {
                "urun_adi": demo.urun_adi,
                "toplam_adet": 0,
                "cihazlar": [],
            }

        grup_sozlugu[anahtar]["toplam_adet"] += 1
        grup_sozlugu[anahtar]["cihazlar"].append(demo)

    demo_gruplari = list(grup_sozlugu.values())

    return render(
        request,
        "stok/demo_cihazlar.html",
        {
            "demo_gruplari": demo_gruplari,
            "arama": arama,
            "aktif_menu": "demo",
        },
    )


@login_required
def demo_ekle(request):
    form = DemoCihazForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Demo cihaz eklendi.")
        return redirect("stok:demo_cihazlar")

    return render(
        request,
        "stok/demo_formu.html",
        {
            "form": form,
            "sayfa_basligi": "YENİ DEMO CİHAZ",
            "aktif_menu": "demo",
        },
    )


@login_required
@require_POST
def demo_sil(request, demo_id):
    demo = get_object_or_404(DemoCihaz, pk=demo_id)
    demo.delete()
    messages.success(request, "Demo cihaz silindi.")
    return redirect("stok:demo_cihazlar")


@login_required
def basit_sayfa(request, baslik, aktif_menu):
    return render(
        request,
        "stok/basit_sayfa.html",
        {
            "baslik": baslik,
            "aktif_menu": aktif_menu,
        },
    )


@login_required
def sterisense(request):
    arama = request.GET.get("q", "").strip()

    kayitlar = SteriSenseKaydi.objects.all().order_by(
        "bayi",
        "firma",
        "urun_adi",
        "seri_no",
    )

    if arama:
        kayitlar = kayitlar.filter(
            Q(bayi__icontains=arama)
            | Q(firma__icontains=arama)
            | Q(urun_adi__icontains=arama)
            | Q(seri_no__icontains=arama)
            | Q(notlar__icontains=arama)
        )

    bayi_sozlugu = {}

    for kayit in kayitlar:
        bayi_anahtari = kayit.bayi.strip().casefold()
        firma_anahtari = kayit.firma.strip().casefold()
        urun_anahtari = kayit.urun_adi.strip().casefold()

        if bayi_anahtari not in bayi_sozlugu:
            bayi_sozlugu[bayi_anahtari] = {
                "bayi": kayit.bayi,
                "toplam_adet": 0,
                "firma_sozlugu": {},
            }

        bayi_grubu = bayi_sozlugu[bayi_anahtari]
        bayi_grubu["toplam_adet"] += 1

        if firma_anahtari not in bayi_grubu["firma_sozlugu"]:
            bayi_grubu["firma_sozlugu"][firma_anahtari] = {
                "firma": kayit.firma,
                "toplam_adet": 0,
                "urun_sozlugu": {},
            }

        firma_grubu = bayi_grubu["firma_sozlugu"][firma_anahtari]
        firma_grubu["toplam_adet"] += 1

        if urun_anahtari not in firma_grubu["urun_sozlugu"]:
            firma_grubu["urun_sozlugu"][urun_anahtari] = {
                "urun_adi": kayit.urun_adi,
                "toplam_adet": 0,
                "kayitlar": [],
            }

        urun_grubu = firma_grubu["urun_sozlugu"][urun_anahtari]
        urun_grubu["toplam_adet"] += 1
        urun_grubu["kayitlar"].append(kayit)

    bayi_gruplari = []

    for bayi_index, bayi_grubu in enumerate(
        bayi_sozlugu.values(),
        start=1,
    ):
        bayi_id = f"bayi-{bayi_index}"
        firmalar = []

        for firma_index, firma_grubu in enumerate(
            bayi_grubu["firma_sozlugu"].values(),
            start=1,
        ):
            firma_id = f"{bayi_id}-firma-{firma_index}"
            urunler = []

            for urun_index, urun_grubu in enumerate(
                firma_grubu["urun_sozlugu"].values(),
                start=1,
            ):
                urun_id = f"{firma_id}-urun-{urun_index}"

                urunler.append(
                    {
                        "id": urun_id,
                        "urun_adi": urun_grubu["urun_adi"],
                        "toplam_adet": urun_grubu["toplam_adet"],
                        "kayitlar": urun_grubu["kayitlar"],
                    }
                )

            firmalar.append(
                {
                    "id": firma_id,
                    "firma": firma_grubu["firma"],
                    "toplam_adet": firma_grubu["toplam_adet"],
                    "urunler": urunler,
                }
            )

        bayi_gruplari.append(
            {
                "id": bayi_id,
                "bayi": bayi_grubu["bayi"],
                "toplam_adet": bayi_grubu["toplam_adet"],
                "firmalar": firmalar,
            }
        )

    return render(
        request,
        "stok/sterisense.html",
        {
            "bayi_gruplari": bayi_gruplari,
            "arama": arama,
            "aktif_menu": "sterisense",
        },
    )


@login_required
def sterisense_ekle(request):
    form = SteriSenseForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()

        messages.success(
            request,
            "SteriSense kaydı eklendi.",
        )

        return redirect("stok:sterisense")

    return render(
        request,
        "stok/sterisense_formu.html",
        {
            "form": form,
            "sayfa_basligi": "YENİ STERISENSE KAYDI",
            "aktif_menu": "sterisense",
        },
    )


@login_required
def sterisense_duzenle(request, kayit_id):
    kayit = get_object_or_404(
        SteriSenseKaydi,
        pk=kayit_id,
    )

    form = SteriSenseForm(
        request.POST or None,
        instance=kayit,
    )

    if request.method == "POST" and form.is_valid():
        form.save()

        messages.success(
            request,
            "SteriSense kaydı güncellendi.",
        )

        return redirect("stok:sterisense")

    return render(
        request,
        "stok/sterisense_formu.html",
        {
            "form": form,
            "sayfa_basligi": "STERISENSE KAYDI DÜZENLE",
            "aktif_menu": "sterisense",
        },
    )


@login_required
@require_POST
def sterisense_sil(request, kayit_id):
    kayit = get_object_or_404(
        SteriSenseKaydi,
        pk=kayit_id,
    )

    kayit.delete()

    messages.success(
        request,
        "SteriSense kaydı silindi.",
    )

    return redirect("stok:sterisense")